from operator import index
from tkinter import * 
import tkinter as tk
from tkinter import filedialog
from tkinter.messagebox import showinfo
from openpyxl import load_workbook
import openpyxl
from tkcalendar import Calendar, DateEntry
from time import strftime
import datetime as dt
import win32com.client 
import pandas as pd 
from pandas import DataFrame as df
import time

main=Tk()
main.title("ExelSheet") 
main.geometry("600x500")
main.config(highlightbackground="black",highlightthickness=2) 

try:
    # Clear Method  

    def clear():
        ini_pro_cost.delete(0,END)
        daily_pro_cost.delete(0,END) 
        penalty.delete(0,END) 
        Bonus.delete(0,END) 
        Budget.delete(0,END) 
        target_dur.delete(0,END)  
    
    # Save Method 
    def Save():
        try:
            a1 = ini_pro_cost.get()
            b1 = daily_pro_cost.get()
            c1 = penalty.get()
            d1 = Bonus.get()
            e1 = Budget.get()
            f1 = target_dur.get()
            g1 = Pro_start_date.get()
            h1 = work_time1.get()
            i1 = work_time2.get()
            work = [(a1,b1,c1,d1,e1,f1,g1,h1,i1)]
            print(work) 
          
            file_path = excel_loc.get() 
            file = openpyxl.load_workbook(file_path)  
            sheet = file['MainProjectInfo'] 
            print(sheet)
            sheet['D2']= a1
            sheet['D3']= b1
            sheet['D4']= c1
            sheet['D5']= d1
            sheet['D6']= e1
            sheet['D7']= f1
            sheet['D8']= g1
            sheet['D9']= h1
            sheet['E9']= i1

            file.save(file_path)
            showinfo("Saved","Data added Successfully")   

        except Exception as x:
            print(x)

    # Browse Path Link 
    def browse():
        print("Browse.............................................")
        file_path = filedialog.askopenfilename(title="any")
        print(file_path) 
        excel_loc.insert(0,file_path) 
    
    # Ms Project Path Link 
    def Path():
        print("Browse.............................................")
        file_path = filedialog.askopenfilename(title= '.mpp')
        print(file_path) 
        ms_project_loc.insert(0,file_path) 
        
    # Ms Project Code Here 
    

    def ms_project():
        Path = ms_project_loc.get()
        file_path = excel_loc.get() 
        file = openpyxl.load_workbook(file_path)  
        sheet1 = file['InputData']
        sheet2 = file['OutputData']
        delete(sheet1) 
        delete(sheet2)
        file.save(file_path) 

        
        try:
            pjApp = win32com.client.Dispatch('MSPRoject.Application')
            pjApp.Visible = 1
            try:
                pjApp.FileOpen(Path) 
                pjproj = pjApp.ActiveProject
                pjTaskList = pjproj.Tasks 
                # taskData = [] 
                # taskData2 = []
               
                # columnNames2 = ['Task name','Start','Duration','Cost']

                
                print(pjproj.Name + "Project  is Open")  
                for task in pjTaskList:
                    # print(task.Name,task.Successors,task.Cost1,task.Cost2,task.Cost3,task.Cost4,task.Cost5,task.Duration1,task.Duration2,task.Duration3,task.Duration4,task.Duration5,task.Start,task.Duration,task.Cost)                     
                    startDate = dt.datetime.strptime(str(task.Start).rstrip('+00:00'),'%Y-%m-%d %H') 
                        # print(startDate) 
                        # # sdate = dt.datetime.strptime(str(startDate),'%Y-%m-%d %H').strftime('%A/%m/%Y %H')
                    
                    Name = task.Name
                    succ = "{}".format(task.Successors) 
                    Suc = succ
                    cost1 = task.Cost1 
                    cost2 = task.Cost2 
                    cost3 = task.Cost3 
                    cost4 = task.Cost4 
                    cost5 = task.Cost5 
                    day = "days" 
                    dur1 = "{}".format(task.Duration1)
                    duration1 = dur1+" " + day
                    dur2 = "{}".format(task.Duration2)
                    duration2 = dur2+" "+day
                    dur3 = "{}".format(task.Duration3)
                    duration3 = dur3+" "+day
                    dur4 = "{}".format(task.Duration4)
                    duration4 = dur4+" "+day
                    dur5 = "{}".format(task.Duration5)
                    duration5 = dur5+ " " +day
                    
                    dur = "{}".format(task.Duration)
                    duration = dur + " "+ day   
                    cost = task.Cost 

                    AllData = [(Name,Suc,cost1,cost2,cost3,cost4,cost5,duration1,duration2,duration3,duration4,duration5)]
                    
                    outdata = [(Name,startDate,duration,cost)]

                    # taskData.append([Name,Suc,cost1,cost2,cost3,cost4,cost5,duration1,duration2,duration3,duration4,duration5]) 
                    # taskData2.append([Name,startDate,duration,cost])
                    try:
                        inputdata(AllData)
                        Output(outdata)  

                    except Exception as e:
                        print(e)      
                    

            except Exception as e:
                print(e) 

            showinfo("Saved","Output and Input Data imported Successfully")
            
            pjApp.Quit() 
            print("File is Closed") 

        except Exception as e:
            print(e) 

    # Input Data Function         
    def inputdata(data):
        try:
            file_path = excel_loc.get() 
            file = openpyxl.load_workbook(file_path)  
            sheet = file['InputData']
            # columnnames1 = ['Task Name','Successors','Cost1','Cost2','Cost3','Cost4','Cost5','Duration1','Duration2','Duration3','Duration4','Duration5']
            # sheet.append(columnnames1) 
            # for dta in data:
            #     sheet.append(dta) 

            # if (sheet.max_row <= 1):
            maxrow = sheet.max_row
            for dta in data:
                for i in range(1,len(dta)+1):
                    cell = sheet.cell(row = maxrow + 1, column = i )
                    # cell = sheet.cell(row = max_row+1,column = i)
                    cell.value = dta[i-1] 
                maxrow += 1
            
            file.save(file_path)
        
        except Exception as e:
            print(e) 
            
    def Output(data):
        try:
            file_path = excel_loc.get() 
            file = openpyxl.load_workbook(file_path)  
            sheet = file['OutputData']

            for dta in data:
                sheet.append(dta) 
            
            file.save(file_path)
        
        except Exception as e:
            print(e) 

    def delete(sheet):
        while(sheet.max_row > 1): 
            sheet.delete_rows(2)
            print("Data Deleted ") 
        return

    # Import OutputData to Ms Project  .........................................

    def browse_Output_excel():
        print("Browse.............................................")
        file_path = filedialog.askopenfilename(title= 'any')
        print(file_path) 
        output_loc.insert(0,file_path)


    def impot_ms_project():
        Path = ms_project_loc.get()
        try:
            pjApp = win32com.client.Dispatch('MSPRoject.Application')
            pjApp.Visible = 1

            try:
                pjApp.FileOpen(Path) 
                pjproj = pjApp.ActiveProject
                pjTaskList = pjproj.Tasks 
                # taskData = [] 
                # taskData2 = []
                
                # columnNames2 = ['Task name','Start','Duration','Cost']

                
                print(pjproj.Name + "Project  is Open")  
                for task in pjTaskList:
                    print(task.Name,task.Successors,task.Cost1,task.Cost2,task.Cost3,task.Cost4,task.Cost5,task.Duration1,task.Duration2,task.Duration3,task.Duration4,task.Duration5,task.Start,task.Duration,task.Cost)


            except Exception as e:
                print(e)

            # pjApp.Quit() 
            print("File is Closed") 

            file_path = output_loc.get()
            file = openpyxl.load_workbook(file_path)
            sheet = file['OutputData']


            max_row = sheet.max_row
            print(max_row) 

            max_col = sheet.max_column
            # print(c_column) 
         
            # print(out_data)
            # dt = pd.read_excel(file_path,sheet_name='OutputData')
            # print(dt)  

        except Exception as e:
            print(e) 
        
        # file = openpyxl.load_workbook(file_path)
        # sheet = file['OutputData']
        
        # # dto = len(dt) 
         
        

        
        

        



    # # OutPut Function 
    # def outputdata(outdata):
    #     try:
    #         file_path = excel_loc.get() 
    #         file = openpyxl.load_workbook(file_path)  
    #         sheet = file['OutputData'] 
    #         print(sheet)
            
    #         maxrow = sheet.max_row
    #         for abc in outdata:
    #             for i in range(1,len(abc)+1):
    #                 cell = sheet.cell(row = maxrow+1 , column = i)
    #                 # cell = sheet.cell(row = max_row+1,column = i)
    #                 cell.value = abc[i-1] 
    #             maxrow += 1
    #         file.save(file_path)
            
    #     except Exception as e:
    #         print(e) 
        
           



    # Design  Part 
    frame1 = LabelFrame(main, text ="Excel Sheets").pack(expand = 'yes', fill = 'both')

    Label(frame1, text="Initial project cost :").place(x=50,y=30) 
    Label(frame1, text="C.U.").place(x=350,y=30)

    Label(frame1, text="Daily project cost : " ).place(x=50,y=60) 
    Label(frame1, text="C.U. /day" ).place(x=350,y=60) 

    Label(frame1, text="Penalty :").place(x=50,y=90) 
    Label(frame1, text="C.U. /day").place(x=350,y=90) 

    Label(frame1, text="Bonus : " ).place(x=50,y=120) 
    Label(frame1, text="C.U. /day" ).place(x=350,y=120) 

    Label(frame1, text="Project Budget :").place(x=50,y=150) 
    Label(frame1, text="$ C.U. /day").place(x=350,y=150) 

    Label(frame1, text="Project target duration :" ).place(x=50,y=180) 
    Label(frame1, text="days" ).place(x=350,y=180) 

    Label(frame1, text="Project start date :").place(x=50,y=210) 

    cal = DateEntry(frame1,locale='en_Us', date_pattern='dd.mm.y', width= 17, background= "magenta3", foreground= "white",bd=2)
    Pro_start_date = cal
    # Pro_start_date.insert(0,cal)
    Pro_start_date.place(x=200,y=210)

    # Label(frame1,cal.pack(pady=10), text= "Choose a Date", background= 'gray61', foreground="white").pack(padx=20,pady=20)
    Label(frame1, text="(MM.DD.YYYY").place(x=350,y=210) 

    Label(frame1, text="Working time : " ).place(x=50,y=240) 
    Label(frame1, text="to").place(x=330,y=240) 
    Label(frame1, text="(HH:MM)").place(x=480,y=240) 

    Label(frame1, text="Excel file location : " ).place(x=50,y=280)

    # Text Field  

    ini_pro_cost = Entry(frame1) 
    ini_pro_cost.place(x=200,y=30) 

    daily_pro_cost = Entry(frame1) 
    daily_pro_cost.place(x=200, y= 60) 

    penalty = Entry(frame1) 
    penalty.place(x=200,y=90) 

    Bonus = Entry(frame1) 
    Bonus.place(x=200,y=120) 

    Budget = Entry(frame1) 
    Budget.place(x=200,y=150) 

    target_dur = Entry(frame1) 
    target_dur.place(x=200,y=180) 

    time = strftime("%H:%M")
    work_time1 = Entry(frame1)
    work_time1.insert(0,time)
    work_time1.place(x=200,y=240) 

    work_time2 = Entry(frame1) 
    work_time2.insert(0,time)
    work_time2.place(x=350,y=240)

    excel_loc = Entry(frame1,width=40)

    excel_loc.place(x=200,y=280) 

    browse_btn = Button(frame1,width="8",height="1",text="Browse" , fg = "blue" , command = browse )
    browse_btn.place(x=470,y=275)   
    
    save_btn = Button(frame1,width="10",height="1",text="Save" , fg = "green" , command = Save )
    save_btn.place(x=210,y=320)

    reset_btn = Button(frame1,width="8",height="1",text="Reset" , fg = "red" , command = clear )
    reset_btn.place(x=320,y=320)

    # For Ms Output Project 
    Label(frame1, text="Export from Ms Project : " ).place(x=50,y=380) 
    ms_project_loc = Entry(frame1,width=30)
    ms_project_loc.place(x=200,y=380) 

    browse_btn = Button(frame1,width="8",height="1",text="Browse" , fg = "blue" , command = Path )
    browse_btn.place(x=395,y=378)

    submit_btn = Button(frame1,width="8",height="1",text="Submit" , fg = "green" , command = ms_project )
    submit_btn.place(x=480,y=378)

    # For Ms Import Project 
    Label(frame1, text="Import into Ms Project : " ).place(x=50,y=430) 
    output_loc= Entry(frame1,width=30)
    output_loc.place(x=200,y=430) 

    browse_excel_btn = Button(frame1,width="8",height="1",text="Browse" , fg = "blue" , command = browse_Output_excel )
    browse_excel_btn.place(x=395,y=428)

    submit_excel = Button(frame1,width="8",height="1",text="Submit" , fg = "green" , command = impot_ms_project )
    submit_excel.place(x=480,y=428)

except Exception as e: 
    print(e) 

main.mainloop()
